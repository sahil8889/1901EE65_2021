# Vaghasiya Sahil
# 1901EE65

from tkinter import *
import tkinter as tk
from tkinter import font
from PIL import Image, ImageTk
from tkinter import filedialog
from tkinter import messagebox
import csv
import pandas as pd
import xlsxwriter
import shutil
import openpyxl
from openpyxl.styles import Border, Side
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
os.system('cls')

master = Tk()

master.title('GUI Based Marksheet Generator')

tk.Label(master, text='Browse for master_roll csv ',font=20).grid(row=0,column=0,pady=10,padx=5)
tk.Label(master, text='Browse for responses csv', font=20).grid(row=2,column=0,pady=10,padx=5)


def file1():
    global df1
    
    filename = filedialog.askopenfilename(initialdir='/sample_input/', title='choose file', filetypes=(('csv file','*.csv'),('all file','*.*')))
    my_label1 = tk.Label(master, text='', font=20)
    my_label1.grid(row=1, column=1)
    my_label1.configure(text=filename)
    df1 = pd.read_csv(filename)
    df1 = df1.set_index('name')

def file2():
    global df

    filename1 = filedialog.askopenfilename(initialdir='/sample_input/', title='choose file', filetypes=(('csv file','*.csv'),('all file','*.*')))
    my_label2 = tk.Label(master, text='', font=20)
    my_label2.grid(row=3, column=1)
    my_label2.configure(text=filename1)
    df = pd.read_csv(filename1)

my_button = tk.Button(master, text='Choose file', command=file1, font=20)
my_button.grid(row=0,column=1,pady=10,padx=5)
my_button = tk.Button(master, text='Choose file', command=file2, font=20)
my_button.grid(row=2,column=1,pady=10,padx=5)

tk.Label(master, text='Marks for correct answer : ', font=20).grid(row=4,column=0,pady=10)
tk.Label(master, text='Marks for incorrect answer : ', font=20).grid(row=5,column=0,pady=10)

e3= tk.Entry(master)
e4 = tk.Entry(master)

e3.grid(row=4,column=1,pady=10,padx=5)
e4.grid(row=5,column=1,pady=10,padx=5)

def reset():
    var= StringVar(master)
    var.set('')
    e3.config(textvariable=var)

def reset1():
    var= StringVar(master)
    var.set('')
    e4.config(textvariable=var)

my_button = tk.Button(master, text='reset', command=reset, font=20)
my_button.grid(row=4,column=2,pady=10,padx=5)

my_button = tk.Button(master, text='reset', command=reset1, font=20)
my_button.grid(row=5,column=2,pady=10,padx=5)


my_label = tk.Label(master, text='', font=20)

def submit():
    for row in range(len(df)):            
        d = {}
        count=0
        d['Roll Number'] = df.iloc[row,6]
        if 'ANSWER' not in d['Roll Number']:
            count+=1
            continue
        else:   
            greet = marksheet(e3.get(),e4.get())
            my_label.config(text=greet)
            messagebox.showinfo('Message info','Roll number wise marksheet generated.')
            break    
    if(count==1):
        messagebox.showinfo('Message info','no roll number with ANSWER is present, Cannot Process!')
        
my_button = tk.Button(master, text='Roll_marksheet', command=submit, font=20)
my_button.grid(row=6,column=0,pady=10,padx=5)

def submit1():
    for row in range(len(df)):            
        d = {}
        count=0
        d['Roll Number'] = df.iloc[row,6]
        if 'ANSWER' not in d['Roll Number']:
            count+=1
            continue
        else:
            greet1 = concise_marksheet(e3.get(),e4.get())
            my_label.config(text=greet1)
            messagebox.showinfo('Message info','Concise marksheet generated')
            break 
    if(count==1):
        messagebox.showinfo('Message info','no roll number with ANSWER is present, Cannot Process!')
        
my_button = tk.Button(master, text='concise marksheet', command=submit1, font=20)
my_button.grid(row=7,column=0,pady=10,padx=5)

def submit2():
    for row in range(len(concise_df)):            
        d = {}
        count=0
        d['Roll Number'] = concise_df.iloc[row,7]
        if 'ANSWER' not in d['Roll Number']:
            count+=1
            continue
        else:
            messagebox.showinfo('Message info','Email will be sent soon.')
            greet2 = email(e3.get(),e4.get())
            my_label.config(text=greet2)
            break 
    if(count==1):
        messagebox.showinfo('Message info','no roll number with ANSWER is present, Cannot Process!')
        
my_button = tk.Button(master, text='send email', command=submit2, font=20)
my_button.grid(row=8,column=0,pady=10,padx=5)

tk.Label(master, text='''Note:  If you want to
generate marksheet again,
then please browse for
responses.csv again.''',font=20).grid(row=9,column=0,pady=10,padx=5)

def marksheet(l, m):
    dict = {}
    for row in range(len(df)):            
        d = {}
        d['Timestamp'] = df.iloc[row,0]
        d['Email address'] = df.iloc[row,1]
        d['Score'] = df.iloc[row,2]
        d['Name'] = df.iloc[row,3]
        d['IITP webmail'] = df.iloc[row,4]
        d['Phone (10 digit only)'] = df.iloc[row,5]
        d['Roll Number'] = df.iloc[row,6]

        d['Answer'] = list(df.iloc[row,7:])

        dict[df.iloc[row,6]] = d 

        ans_key = dict['ANSWER']['Answer'] 

        for x,y in dict.items():
            correct = 0
            incorrect = 0
            na = 0
            for j in range (len(ans_key)):
                if(str(y['Answer'][j])=='nan'):
                    na+=1
                elif(y['Answer'][j]== ans_key[j]):
                    correct +=1
                else:
                    incorrect+=1

            dict[x]['correct'] = correct
            dict[x]['incorrect'] = incorrect
            dict[x]['na'] = na

        wb = openpyxl.Workbook()
        wb.create_sheet(index = 0, title = 'quiz')
        sheet = wb['quiz']
        img = openpyxl.drawing.image.Image('IITP.jpeg')
        img.anchor = 'A1'
        sheet.add_image(img)

        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        sheet.column_dimensions['A'].width = 16.89
        sheet.column_dimensions['B'].width = 16.89
        sheet.column_dimensions['C'].width = 16.89
        sheet.column_dimensions['D'].width = 16.89
        sheet.column_dimensions['E'].width = 16.89

        sheet.row_dimensions[5].height = 22.8
        sheet. merge_cells('A5:E5')
        sheet['A5'].font = openpyxl.styles.Font(name = 'Century',size=18, bold=True,underline='single')
        sheet['A5'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
        sheet['A5'] = 'Mark Sheet'

        sheet['A6'] = 'Name:'
        sheet['A6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
        sheet['A6'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

        sheet['B6'] = d['Name']
        sheet['B6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['B6'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

        sheet['D6'] = 'Exam:'
        sheet['D6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
        sheet['D6'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

        sheet['E6'] = 'quiz'
        sheet['E6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['E6'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

        sheet['A7'] = 'Roll Number:'
        sheet['A7'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
        sheet['A7'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

        sheet['B7'] = d['Roll Number']
        sheet['B7'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['B7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

        sheet['B9'] = 'Right'
        sheet['B9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['B9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['C9'] = 'Wrong'
        sheet['C9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['C9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['D9'] = 'Not Attempt'
        sheet['D9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['D9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['E9'] = 'Max'
        sheet['E9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['E9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['A10'] = 'No.'
        sheet['A10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['A10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['A11'] = 'Marking'
        sheet['A11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['A11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['A12'] = 'Total'
        sheet['A12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['A12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['B10'] = correct
        sheet['B10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
        sheet['B10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['B11'] = l
        sheet['B11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
        sheet['B11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['B12'] = (correct * float(l))
        sheet['B12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
        sheet['B12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['C10'] = incorrect
        sheet['C10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
        sheet['C10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['C11'] = m
        sheet['C11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
        sheet['C11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['C12'] = (incorrect * float(m))
        sheet['C12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
        sheet['C12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['D10'] = (len(ans_key)-(correct)-(incorrect))
        sheet['D10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
        sheet['D10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['D11'] = '0'
        sheet['D11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
        sheet['D11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['E10'] = len(ans_key)
        sheet['E10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
        sheet['E10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['E12'] = str((correct * float(l))+(incorrect * float(m)))+'/'+str(float(l) * len(ans_key))
        sheet['E12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
        sheet['E12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

        sheet['A15'] = 'Student Ans'
        sheet['A15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['A15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
        sheet['A15'].border = border

        sheet['D15'] = 'Student Ans'
        sheet['D15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['D15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
        sheet['D15'].border = border

        sheet['B15'] = 'Correct Ans'
        sheet['B15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['B15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
        sheet['B15'].border = border

        sheet['E15'] = 'Correct Ans'
        sheet['E15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
        sheet['E15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
        sheet['E15'].border = border

        for r in range(9,13):
            for c in range(1,6):
                sheet.cell(row = r , column = c).border = border

        for i in range(len(ans_key)):
            if(i+16>40):
                sheet['E'+str(i-9)].border = border
                
                sheet['E'+str(i-9)] = ans_key[i]
                sheet['E'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
                sheet['E'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
            else:   
                sheet['B'+str(i+16)].border = border
                
                sheet['B'+str(i+16)] = ans_key[i]
                sheet['B'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
                sheet['B'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                
        for i in range(len(ans_key)):
            if(i+16>40):
                sheet['D'+str(i-9)].border = border
                
                if(str(dict[d['Roll Number']]['Answer'][i]) == 'nan'):
                    pass
                elif(dict[d['Roll Number']]['Answer'][i] == ans_key[i]):
                    sheet['D'+str(i-9)] = dict[d['Roll Number']]['Answer'][i]
                    sheet['D'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                    sheet['D'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                else:
                    sheet['D'+str(i-9)] = dict[d['Roll Number']]['Answer'][i]
                    sheet['D'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                    sheet['D'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')            
                    
            else:
                sheet['A'+str(i+16)].border = border
                    
                if(str(dict[d['Roll Number']]['Answer'][i]) == 'nan'):
                    pass
                elif(dict[d['Roll Number']]['Answer'][i] == ans_key[i]):
                    sheet['A'+str(i+16)] = dict[d['Roll Number']]['Answer'][i]
                    sheet['A'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                    sheet['A'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                else:
                    sheet['A'+str(i+16)] = dict[d['Roll Number']]['Answer'][i]
                    sheet['A'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                    sheet['A'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom') 
           
        sheetDelete = wb["Sheet"]
        wb.remove(sheetDelete)
        path2 = 'output/' + d['Roll Number'] + '.xlsx'
        wb.save(path2)

def concise_marksheet(l, m):
    global concise_df
    concise_df = df
    dict = {}
    for row in range(len(concise_df)):            
        d = {}
        d['Answer'] = list(concise_df.iloc[row,7:])

        dict[concise_df.iloc[row,6]] = d 

        ans_key = dict['ANSWER']['Answer'] 

        for x,y in dict.items():
            correct = 0
            incorrect = 0
            na = 0
            for j in range (len(ans_key)):
                if(str(y['Answer'][j])=='nan'):
                    na+=1
                elif(y['Answer'][j]== ans_key[j]):
                    correct +=1
                else:
                    incorrect+=1

            dict[x]['correct'] = correct
            dict[x]['incorrect'] = incorrect
            dict[x]['na'] = na

    concise_df.insert(6,'Score_After_Negative','any')
    concise_df.insert(36, 'statusAns','any')
    for k in range(len(concise_df)):
            concise_df.iloc[k,6] = str(dict[df.iloc[k,7]]['correct']*float(l) + dict[df.iloc[k,7]]['incorrect']*float(m)) + '/' + str(float(l)*len(ans_key))
            concise_df.iloc[k,36] = '[' + str(dict[df.iloc[k,7]]['correct']) + ',' + str(dict[df.iloc[k,7]]['incorrect']) + ',' + str(dict[df.iloc[k,7]]['na']) + ']'
                
    df4= pd.DataFrame(concise_df)
    path2 = 'output/concise marksheet.csv'
    df4.to_csv(path2, index=False)

def email(l, m):
    for row in range(len(concise_df)):            
        d = {}
        d['Email address'] = concise_df.iloc[row,1]
        d['IITP webmail'] = concise_df.iloc[row,4]
        d['Roll Number'] = concise_df.iloc[row,7]

        email_user = 'pythoncs384sahilee65@gmail.com'
        email_password = 'pythonCS384sahilEE65'
        email_send = d['Email address']
        email_send1 = d['IITP webmail']

        subject = 'CS384 - Quiz Exam - with Negative'

        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = email_send
        msg['Subject'] = subject

        msg1 = MIMEMultipart()
        msg1['From'] = email_user
        msg1['To'] = email_send1
        msg1['Subject'] = subject

        body = """"Dear Student,
                
                Quiz Exam marks are attached for reference."""
        msg.attach(MIMEText(body,'plain'))
        msg1.attach(MIMEText(body,'plain'))

        filename = d['Roll Number'] + '.xlsx'
        attachment  = open('output/'+filename,'rb')
        
        part = MIMEBase('application','octet-stream')
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',"attachment; filename= "+filename)

        msg.attach(part)
        msg1.attach(part)
        text = msg.as_string()
        text1 = msg1.as_string()
        server = smtplib.SMTP('smtp.gmail.com',587)
        server.starttls()
        server.login(email_user,email_password)
        server.sendmail(email_user,email_send,text)
        server.sendmail(email_user,email_send1,text1)
        server.quit()

master.mainloop()