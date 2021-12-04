#from _typeshed import Self
import pandas as pd
from fpdf import FPDF
import os
import csv
import openpyxl
import datetime

class PDF(FPDF):
    def lines(self):
        self.rect(5.0, 5.0, 287.0, 360.0)
        self.line(5, 69, 292, 69)
        self.line(5, 150, 292, 150)
        self.line(5, 230, 292, 230)
        self.line(5, 310, 292, 310)
        self.line(97.33, 69, 97.33, 310)
        self.line(194.66, 69, 194.66, 310)
        self.rect(45.0, 5.0, 207, 35)
        self.rect(5.0, 40.0, 40, 0)
        self.rect(252.0, 40.0, 40, 0)
        self.rect(45.0, 45, 217, 19)
        #self.rect(50.0,54.5,197,0)
    
    def header(self):
        self.image(r'symbol.jpeg', 10,7,30,25)
        self.image(r'symbol.jpeg', 257,7,30,25)
        self.image(r'interim.jpeg', 9,34,33,5)
        self.image(r'interim.jpeg', 256,34,33,5)
        self.image(r'iitpatna(hindi).jpeg', 60,7,180,12)
        self.image(r'iitpatna.jpeg', 53,21,190,12)
        self.image(r'transcript.jpeg',123.5,33,60,6)
        self.image(r'roll no.jpeg', 47,47,15,5.5)
        self.image(r'name.jpeg', 122,47,15,5.5)
        self.image(r'year.jpeg', 200,47,35,5.5)
        self.image(r'programme.jpeg', 47,56.5,20,5.5)
        self.set_font('helvetica', '', 10)
        self.text(69, 60.5, 'Bechelor of Technolgy')
        self.image(r'course.jpeg', 122,56.5,18,5.5)
        self.image(r'stamp.jpeg', 122,315,40,40)
        self.set_font('helvetica', '', 10)
        self.text(7, 335, 'Date of Issue : ')
        self.set_font('helvetica', '', 10)
        self.text(222, 340, 'Assistant Register (academic) ')
        t= datetime.datetime.now()
        self.set_font('helvetica', '', 10)
        self.text(7, 355, f'{t.strftime}')


    def table(self):
        with open('names-roll.csv', 'r') as csvfile:
            csvreader = csv.DictReader(csvfile)
            name = {}
            for row in csvreader:
                #r = row['Roll']
                name[row['Roll']] = row['Name']
        #if roll_no in r:
            with open('subjects_master.csv', 'r') as csvfile:
                csvreader = csv.DictReader(csvfile)
                subject = {}
                for row in csvreader:
                    list = []
                    list.append(row['subname'])
                    list.append(row['ltp'])
                    subject[row['subno']] = list

                    with open('grades.csv', 'r') as csvfile:
                        csvreader = csv.DictReader(csvfile)
                        grades = {}
                        for row in csvreader:
                            list = []
                            list.append(row['Roll'])
                            list.append(row['Sem'])
                            list.append(row['SubCode'])
                            list.append(row['Credit'])
                            list.append(row['Grade'])
                            list.append(row['Sub_Type'])
                            if(grades.get(row['Roll']) == None):
                                grades[row['Roll']] = [list]
                            else:
                                grades[row['Roll']].append(list)

                d = {}
                for roll in grades:
                    d1 = {}
                    for row in grades[roll]:
                        list = []
                        list.append(row[2])
                        list.append(subject[row[2]][0])
                        list.append(subject[row[2]][1])
                        list.append(row[3])
                        list.append(row[5])
                        list.append(row[4])
                        if(d1.get(row[1]) == None):
                            d1[row[1]] = [list]
                        else:  
                            d1[row[1]].append(list)
                    d[roll] = d1
                for roll in grades: 
                    GNE = {'AA':10,'AB':9,'BB':8,' BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,'DD*':4,'F*':0}
                    semester = []
                    spi = []
                    cpi = []
                    cred_taken = []
                    total_cred_taken = []
                    total_cred_sum = 0
                    cpi_sum = 0
                    for i in d[roll]:
                        semester.append(i)
                        spi_sum = 0
                        cred_sum = 0
                        for row in d[roll][i]:
                            marks = float(GNE[row[5]])
                            cred = float(row[3])
                            spi_sum += marks*cred
                            cred_sum += cred
                        total_cred_sum += cred_sum
                        cpi_sum += (spi_sum/cred_sum)*cred_sum
                        spi.append(round(spi_sum/cred_sum, 2))
                        cpi.append(round(cpi_sum/total_cred_sum, 2))
                        cred_taken.append(cred_sum)
                        total_cred_taken.append(total_cred_sum)
                self.set_font('helvetica', 'B', 6)
                self.set_xy(6, 127)
                self.cell(75, 4,('Credits Taken: '+str(cred_taken[0])+'    Credits Cleared: '+str(cred_taken[0])+'    SPI: '+str(spi[0]) +'    CPI: '+str(cpi[0])),border= 1,align= 'C')

                self.set_font('helvetica', 'B', 6)
                self.set_xy(100.33, 127)
                self.cell(75, 4,('Credits Taken: '+str(cred_taken[1])+'    Credits Cleared: '+str(cred_taken[1])+'    SPI: '+str(spi[1]) +'    CPI: '+str(cpi[1])),border= 1,align= 'C')

                self.set_font('helvetica', 'B', 6)
                self.set_xy(196.66, 127)
                self.cell(75, 4,('Credits Taken: '+str(cred_taken[2])+'    Credits Cleared: '+str(cred_taken[2])+'    SPI: '+str(spi[2]) +'    CPI: '+str(cpi[2])),border= 1,align= 'C')

                self.set_font('helvetica', 'B', 6)
                self.set_xy(6, 214)
                self.cell(75, 4,('Credits Taken: '+str(cred_taken[3])+'    Credits Cleared: '+str(cred_taken[3])+'    SPI: '+str(spi[3]) +'    CPI: '+str(cpi[3])),border= 1,align= 'C')

                self.set_font('helvetica', 'B', 6)
                self.set_xy(100.33, 214)
                self.cell(75, 4,('Credits Taken: '+str(cred_taken[4])+'    Credits Cleared: '+str(cred_taken[4])+'    SPI: '+str(spi[4]) +'    CPI: '+str(cpi[4])),border= 1,align= 'C')

                self.set_font('helvetica', 'B', 6)
                self.set_xy(196.66, 214)
                self.cell(75, 4,('Credits Taken: '+str(cred_taken[5])+'    Credits Cleared: '+str(cred_taken[5])+'    SPI: '+str(spi[5]) +'    CPI: '+str(cpi[5])),border= 1,align= 'C')

                self.set_font('helvetica', 'B', 6)
                self.set_xy(6, 283)
                self.cell(75, 4,('Credits Taken: '+str(cred_taken[6])+'    Credits Cleared: '+str(cred_taken[6])+'    SPI: '+str(spi[6]) +'    CPI: '+str(cpi[6])),border= 1,align= 'C')

                self.set_font('helvetica', 'B', 6)
                self.set_xy(98.33, 283)
                self.cell(75, 4,('Credits Taken: '+str(cred_taken[0])+'    Credits Cleared: '+str(cred_taken[0])+'    SPI: '+str(spi[0]) +'    CPI: '+str(cpi[0])),border= 1,align= 'C')

    def xy(self):
        with open('names-roll.csv', 'r') as csvfile:
            csvreader = csv.DictReader(csvfile)
            name = {}
            for row in csvreader:
                #r = row['Roll']
                name[row['Roll']] = row['Name']
        #roll_no = '04010ME14'
        #if roll_no == r:
            with open('subjects_master.csv', 'r') as csvfile:
                csvreader = csv.DictReader(csvfile)
                subject = {}
                for row in csvreader:
                    list = []
                    list.append(row['subname'])
                    list.append(row['ltp'])
                    subject[row['subno']] = list

                with open('grades.csv', 'r') as csvfile:
                    csvreader = csv.DictReader(csvfile)
                    grades = {}
                    for row in csvreader:
                        list = []
                        list.append(row['Roll'])
                        list.append(row['Sem'])
                        list.append(row['SubCode'])
                        list.append(row['Credit'])
                        list.append(row['Grade'])
                        list.append(row['Sub_Type'])
                        if(grades.get(row['Roll']) == None):
                            grades[row['Roll']] = [list]
                        else:
                            grades[row['Roll']].append(list)

            d = {}
            for roll in grades:
                d1 = {}
                #for i in d[roll]:
                for row in grades[roll]:
                        list = []
                        l1 = ['SubCode', 'Subject Name', 'L-T-P', 'CRD', 'SubType', 'GRD']
                        list.append(row[2])
                        list.append(subject[row[2]][0])
                        list.append(subject[row[2]][1])
                        list.append(row[3])
                        list.append(row[5])
                        list.append(row[4])
                        list1 = list
                        if(d1.get(row[1]) == None):
                            d1[row[1]] = [l1] + [list1]
                        else:
                            d1[row[1]].append(list)
                d[roll] = d1
            
            self.set_font('helvetica', 'BU', 10)
            self.text(6, 72, 'semester 1')
            self.y = 75
            for row in d[roll]['1']:
                self.x = self.x - 4
                self.set_font('helvetica', 'B', 6)
                self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(5)

            self.set_font('helvetica', 'BU', 10)
            self.text(98.33, 72, 'semester 2')
            self.y = 75
            for row in d[roll]['2']:
                self.x = self.x + 94.33 - 4
                self.set_font('helvetica', 'B', 6)
                self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(5)

            self.set_font('helvetica', 'BU', 10)
            self.text(194.66, 72, 'semester 3')
            self.y = 75
            for row in d[roll]['3']:
                self.x = self.x + 190.66 - 4
                self.set_font('helvetica', 'B', 6)
                self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(5)
            
            self.set_font('helvetica', 'BU', 10)
            self.text(6, 153, 'semester 4')
            self.y = 156
            for row in d[roll]['4']:
                self.x = self.x - 4
                self.set_font('helvetica', 'B', 6)
                self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(5)
                
            self.set_font('helvetica', 'BU', 10)
            self.text(98.33, 153, 'semester 5')
            self.y = 156
            for row in d[roll]['5']:
                self.x = self.x + 94.33 - 4
                self.set_font('helvetica', 'B', 6)
                self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(5)

            self.set_font('helvetica', 'BU', 10)
            self.text(194.66, 153, 'semester 6')
            self.y = 156
            for row in d[roll]['6']:
                self.x = self.x + 190.66 - 4
                self.set_font('helvetica', 'B', 6)
                self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(5)

            self.set_font('helvetica', 'BU', 10)
            self.text(6, 233, 'semester 7')
            self.y = 236
            for row in d[roll]['7']:
                self.x = self.x - 4
                self.set_font('helvetica', 'B', 6)
                self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(5)

            self.set_font('helvetica', 'BU', 10)
            self.text(98.33, 233, 'semester 8')
            self.y = 236
            for row in d[roll]['8']:
                self.x = self.x + 94.33 - 4
                self.set_font('helvetica', 'B', 6)
                self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(5)
    
    def final(self):
        global roll_no
        roll_no = '0401CS02'
        self.add_page()
        self.lines()
        self.header()
        self.xy()
        self.table()

pdf = PDF('P', 'mm', 'A3')
#pdf.add_page()
# pdf.lines()
# pdf.header()



pdf.output('a.pdf','F')