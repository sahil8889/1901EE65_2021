# Vaghasiya Sahil
#1901EE65

from tkinter import *
import tkinter as tk
from tkinter import font
from PIL import Image, ImageTk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
from fpdf import FPDF
import os
import csv
import openpyxl
import re
import datetime

filename = None
filename1 = None

class PDF(FPDF):
    def lines(self):
        self.rect(5.0, 5.0, 287.0, 360.0)
        self.line(5, 69, 292, 69)
        self.line(5, 150, 292, 150)
        self.line(5, 230, 292, 230)
        self.line(5, 310, 292, 310)
        self.line(97.33, 69, 97.33, 310)
        self.line(194.66, 69, 194.66, 310)
        self.rect(45.0, 5.0, 207, 35)
        self.rect(5.0, 40.0, 40, 0)
        self.rect(252.0, 40.0, 40, 0)
        self.rect(45.0, 45, 217, 19)
    
    def header(self):
        self.image(r'images\\symbol.jpeg', 10,7,30,25)
        self.image(r'images\\symbol.jpeg', 257,7,30,25)
        self.image(r'images\\interim.jpeg', 9,34,33,5)
        self.image(r'images\\interim.jpeg', 256,34,33,5)
        self.image(r'images\\iitpatna(hindi).jpeg', 60,7,180,12)
        self.image(r'images\\iitpatna.jpeg', 53,21,190,12)
        self.image(r'images\\transcript.jpeg',123.5,33,60,6)
        self.image(r'images\\roll no.jpeg', 47,47,15,5.5)
        self.image(r'images\\name.jpeg', 122,47,15,5.5)
        self.image(r'images\\year.jpeg', 200,47,35,5.5)
        self.image(r'images\\programme.jpeg', 47,56.5,20,5.5)
        self.set_font('helvetica', '', 12)
        self.text(69, 60.5, 'Bachelor of Technolgy')
        self.image(r'images\\course.jpeg', 122,56.5,18,5.5)
        if filename != None:
            self.image(fr'{filename}', 122,315,40,40)
        if filename1 != None:
            self.image(fr'{filename1}',230,320,20,15)
        self.set_font('helvetica', '', 10)
        self.text(7, 335, 'Date of Issue : ')
        self.set_font('helvetica', '', 10)
        self.text(222, 340, 'Assistant Registrar (academic) ')
        t= datetime.datetime.now()
        self.set_font('helvetica', '', 10)
        self.text(32, 335, f'{t}')
        
    def table(self, roll_no):
        with open('sample_input/names-roll.csv', 'r') as csvfile:
            csvreader = csv.DictReader(csvfile)
            name = {}
            for row in csvreader:
                r = row['Roll']
                name[row['Roll']] = row['Name']
                if roll_no == r:
                    with open('sample_input/subjects_master.csv', 'r') as csvfile:
                        csvreader = csv.DictReader(csvfile)
                        subject = {}
                        for row in csvreader:
                            list = []
                            list.append(row['subname'])
                            list.append(row['ltp'])
                            subject[row['subno']] = list

                            with open('sample_input/grades.csv', 'r') as csvfile:
                                csvreader = csv.DictReader(csvfile)
                                grades = {}
                                for row1 in csvreader:
                                    list = []
                                    list.append(row1['Roll'])
                                    list.append(row1['Sem'])
                                    list.append(row1['SubCode'])
                                    list.append(row1['Credit'])
                                    list.append(row1['Grade'])
                                    list.append(row1['Sub_Type'])
                                    if(grades.get(row1['Roll']) == None):
                                        grades[row1['Roll']] = [list]
                                    else:
                                        grades[row1['Roll']].append(list)

                        d = {}
                        for roll in grades:
                            d1 = {}
                            for row2 in grades[r]:
                                list = []
                                list.append(row2[2])
                                list.append(subject[row2[2]][0])
                                list.append(subject[row2[2]][1])
                                list.append(row2[3])
                                list.append(row2[5])
                                list.append(row2[4])
                                if(d1.get(row2[1]) == None):
                                    d1[row2[1]] = [list]
                                else:  
                                    d1[row2[1]].append(list)
                            d[roll] = d1
                        for roll in grades: 
                            GNE = {'AA':10,'AB':9,'BB':8,' BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,'DD*':4,'F*':0}
                            semester = []
                            spi = []
                            cpi = []
                            cred_taken = []
                            total_cred_taken = []
                            cred_taken_cleared = []
                            total_cred_taken_creared = []
                            cpi_sum = 0
                            total_cred_sum = 0
                            total_cred_cleared = 0
                            for i in d[roll]:
                                semester.append(i)
                                spi_sum = 0
                                cred_sum = 0
                                cred_cleared = 0
                                for row in d[roll][i]:
                                    marks = float(GNE[row[5]])
                                    cred = float(row[3])
                                    spi_sum += marks*cred
                                    cred_sum += cred
                                    if row[5] == 'DD*' or row[5] == 'F' or row[5] == 'F*' or row[5] == 'I':
                                        pass
                                    else:
                                        cred_cleared += cred
                                total_cred_sum += cred_sum
                                total_cred_cleared += cred_cleared
                                cpi_sum += (spi_sum/cred_sum)*cred_sum
                                spi.append(round(spi_sum/cred_sum, 2))
                                cpi.append(round(cpi_sum/total_cred_sum, 2))
                                cred_taken.append(cred_sum)
                                cred_taken_cleared.append(cred_cleared)
                                total_cred_taken_creared.append(total_cred_taken_creared)
                                total_cred_taken.append(total_cred_sum)
                        self.set_font('helvetica', 'B', 6)
                        self.set_xy(6, 127)
                        self.cell(75, 4,('Credits Taken: '+str(cred_taken[0])+'    Credits Cleared: '+str(cred_taken_cleared[0])+'    SPI: '+str(spi[0]) +'    CPI: '+str(cpi[0])),border= 1,align= 'C')

                        self.set_font('helvetica', 'B', 6)
                        self.set_xy(100.33, 127)
                        self.cell(75, 4,('Credits Taken: '+str(cred_taken[1])+'    Credits Cleared: '+str(cred_taken_cleared[1])+'    SPI: '+str(spi[1]) +'    CPI: '+str(cpi[1])),border= 1,align= 'C')

                        self.set_font('helvetica', 'B', 6)
                        self.set_xy(196.66, 127)
                        self.cell(75, 4,('Credits Taken: '+str(cred_taken[2])+'    Credits Cleared: '+str(cred_taken_cleared[2])+'    SPI: '+str(spi[2]) +'    CPI: '+str(cpi[2])),border= 1,align= 'C')

                        self.set_font('helvetica', 'B', 6)
                        self.set_xy(6, 214)
                        self.cell(75, 4,('Credits Taken: '+str(cred_taken[3])+'    Credits Cleared: '+str(cred_taken_cleared[3])+'    SPI: '+str(spi[3]) +'    CPI: '+str(cpi[3])),border= 1,align= 'C')

                        self.set_font('helvetica', 'B', 6)
                        self.set_xy(100.33, 214)
                        self.cell(75, 4,('Credits Taken: '+str(cred_taken[4])+'    Credits Cleared: '+str(cred_taken_cleared[4])+'    SPI: '+str(spi[4]) +'    CPI: '+str(cpi[4])),border= 1,align= 'C')

                        self.set_font('helvetica', 'B', 6)
                        self.set_xy(196.66, 214)
                        self.cell(75, 4,('Credits Taken: '+str(cred_taken[5])+'    Credits Cleared: '+str(cred_taken_cleared[5])+'    SPI: '+str(spi[5]) +'    CPI: '+str(cpi[5])),border= 1,align= 'C')

                        self.set_font('helvetica', 'B', 6)
                        self.set_xy(6, 283)
                        self.cell(75, 4,('Credits Taken: '+str(cred_taken[6])+'    Credits Cleared: '+str(cred_taken_cleared[6])+'    SPI: '+str(spi[6]) +'    CPI: '+str(cpi[6])),border= 1,align= 'C')

                        self.set_font('helvetica', 'B', 6)
                        self.set_xy(98.33, 283)
                        self.cell(75, 4,('Credits Taken: '+str(cred_taken[7])+'    Credits Cleared: '+str(cred_taken_cleared[7])+'    SPI: '+str(spi[7]) +'    CPI: '+str(cpi[7])),border= 1,align= 'C')

                        if roll_no == '0401ME11':
                            self.set_font('helvetica', 'B', 6)
                            self.set_xy(196.66, 283)
                            self.cell(75, 4,('Credits Taken: '+str(cred_taken[8])+'    Credits Cleared: '+str(cred_taken_cleared[8])+'    SPI: '+str(spi[8]) +'    CPI: '+str(cpi[8])),border= 1,align= 'C')

        roll1 = str(roll_no)
        with open('sample_input/names-roll.csv', 'r') as csvfile:
            csvreader = csv.reader(csvfile)
            for row in csvreader:
                if roll_no != row[0]:
                    continue
                else:
                    name1 = row[1]
                    break
        s = roll1[4:6]
        e = roll1[0:2]
        name = str(name1)

        e = '20' + e
        if s == 'CS':
            s = 'Computer Science and Engineering'
        if s == 'EE':
            s='Electrical and Electronics Engineering'
        if s == 'ME':
            s = 'Mechanical Engineering'

        self.set_font('helvetica', '', 12)
        self.text(69, 51.5, roll1)
        self.set_font('helvetica', '', 12)
        self.text(142, 51.5, name)
        self.set_font('helvetica', '', 12)
        self.text(142, 60.5, s)
        self.set_font('helvetica', '', 12)
        self.text(237, 51.5, e)

    def xy(self, roll_no):
        with open('sample_input/names-roll.csv', 'r') as csvfile:
            csvreader = csv.DictReader(csvfile)
            name = {}
            for row in csvreader:
                r = row['Roll']
                name[row['Roll']] = row['Name']
                if roll_no == r:
                    with open('sample_input/subjects_master.csv', 'r') as csvfile:
                        csvreader = csv.DictReader(csvfile)
                        subject = {}
                        for row in csvreader:
                            list = []
                            list.append(row['subname'])
                            list.append(row['ltp'])
                            subject[row['subno']] = list

                        with open('sample_input/grades.csv', 'r') as csvfile:
                            csvreader = csv.DictReader(csvfile)
                            grades = {}
                            for row in csvreader:
                                list = []
                                list.append(row['Roll'])
                                list.append(row['Sem'])
                                list.append(row['SubCode'])
                                list.append(row['Credit'])
                                list.append(row['Grade'])
                                list.append(row['Sub_Type'])
                                if(grades.get(row['Roll']) == None):
                                    grades[row['Roll']] = [list]
                                else:
                                    grades[row['Roll']].append(list)

                    d = {}
                    for roll in grades:
                        d1 = {}
                        for row in grades[r]:
                                list = []
                                l1 = ['SubCode', 'Subject Name', 'L-T-P', 'CRD', 'SubType', 'GRD']
                                list.append(row[2])
                                list.append(subject[row[2]][0])
                                list.append(subject[row[2]][1])
                                list.append(row[3])
                                list.append(row[5])
                                list.append(row[4])
                                list1 = list
                                if(d1.get(row[1]) == None):
                                    d1[row[1]] = [l1] + [list1]
                                else:
                                    d1[row[1]].append(list)
                        d[roll] = d1
                    
                    self.set_font('helvetica', 'BU', 10)
                    self.text(6, 72, 'semester 1')
                    self.y = 75
                    for row in d[roll]['1']:
                        self.x = self.x - 4
                        self.set_font('helvetica', 'B', 6)
                        self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.ln(5)

                    self.set_font('helvetica', 'BU', 10)
                    self.text(98.33, 72, 'semester 2')
                    self.y = 75
                    for row in d[roll]['2']:
                        self.x = self.x + 94.33 - 4
                        self.set_font('helvetica', 'B', 6)
                        self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.ln(5)

                    self.set_font('helvetica', 'BU', 10)
                    self.text(194.66, 72, 'semester 3')
                    self.y = 75
                    for row in d[roll]['3']:
                        self.x = self.x + 190.66 - 4
                        self.set_font('helvetica', 'B', 6)
                        self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.ln(5)
                    
                    self.set_font('helvetica', 'BU', 10)
                    self.text(6, 153, 'semester 4')
                    self.y = 156
                    for row in d[roll_no]['4']:
                        self.x = self.x - 4
                        self.set_font('helvetica', 'B', 6)
                        self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.ln(5)
                        
                    self.set_font('helvetica', 'BU', 10)
                    self.text(98.33, 153, 'semester 5')
                    self.y = 156
                    for row in d[roll]['5']:
                        self.x = self.x + 94.33 - 4
                        self.set_font('helvetica', 'B', 6)
                        self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.ln(5)

                    self.set_font('helvetica', 'BU', 10)
                    self.text(194.66, 153, 'semester 6')
                    self.y = 156
                    for row in d[roll]['6']:
                        self.x = self.x + 190.66 - 4
                        self.set_font('helvetica', 'B', 6)
                        self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.ln(5)

                    self.set_font('helvetica', 'BU', 10)
                    self.text(6, 233, 'semester 7')
                    self.y = 236
                    for row in d[roll]['7']:
                        self.x = self.x - 4
                        self.set_font('helvetica', 'B', 6)
                        self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.ln(5)

                    self.set_font('helvetica', 'BU', 10)
                    self.text(98.33, 233, 'semester 8')
                    self.y = 236
                    for row in d[roll]['8']:
                        self.x = self.x + 94.33 - 4
                        self.set_font('helvetica', 'B', 6)
                        self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                        self.ln(5)
                    if roll_no == '0401ME11':
                        self.set_font('helvetica', 'BU', 10)
                        self.text(194.33, 233, 'semester 10')
                        self.y = 236
                        for row in d[roll]['10']:
                            self.x = self.x + 190.33 - 4
                            self.set_font('helvetica', 'B', 6)
                            self.multi_cell(15, 5, row[0], border=1, align='C', ln=3, max_line_height=self.font_size)
                            self.multi_cell(49, 5, row[1], border=1, align='C', ln=3, max_line_height=self.font_size)
                            self.multi_cell(10, 5, row[2], border=1, align='C', ln=3, max_line_height=self.font_size)
                            self.multi_cell(8, 5, row[3], border=1, align='C', ln=3, max_line_height=self.font_size)
                            self.multi_cell(8, 5, row[5], border=1, align='C', ln=3, max_line_height=self.font_size)
                            self.ln(5)

    def final(self, roll_no):
        self = PDF('P', 'mm', 'A3')
        self.add_page()
        self.lines()
        self.header()
        self.xy(roll_no)
        self.table(roll_no)
        self.output(f'output/{roll_no}.pdf','F')


pdf = PDF('P', 'mm', 'A3')

def execute_range(str1, str2):
    s1 = int(str1[6:8])
    s2 = int(str2[6:8])
    l= []
    st = str1[:6]
    for x in range(s1,s2+1):
        with open('sample_input/names-roll.csv','r') as f:
            reader = csv.reader(f)
            c=0
            for the_row in reader:
                x = str(x)
                if the_row[0] ==st + x.zfill(2):
                    c+=1
                    pdf.final(the_row[0])
                    continue
        if c==0:
            n = st + str(x)
            l.append(n)
    return l

def check_input(str1, str2):
    for x in range(0,6):
        if str1[x]!=str2[x]:
            return False
        else:
            continue
    pattern =re.compile(r'[0-9]{4}[A-Z]{2}[0-9]{2}')
    if re.match(pattern, str1) and re.match(pattern, str2):
        return True
    return False

def execute_all():
    with open('sample_input/names-roll.csv','r') as f:
        reader = csv.reader(f)
        for the_row in reader:
            pdf.final(the_row[0])
            continue
    return

master = tk.Tk()

master.title('GUI Based Transcript Generator ')

tk.Label(master, text='From Roll No : ', font=20).grid(row=0,column=0,pady=10,padx=5)
tk.Label(master, text='to Roll No : ', font=20).grid(row=1,column=0,pady=10,padx=5)

e3 = tk.Entry(master)
e4 = tk.Entry(master)

e3.grid(row=0,column=1,pady=10,padx=5)
e4.grid(row=1,column=1,pady=10,padx=5)

def reset():
    var= StringVar(master)
    var.set('')
    e3.config(textvariable=var)

def reset1():
    var= StringVar(master)
    var.set('')
    e4.config(textvariable=var)

my_button = tk.Button(master, text='reset', command=reset, font=20)
my_button.grid(row=0,column=2,pady=10,padx=5)

my_button = tk.Button(master, text='reset', command=reset1, font=20)
my_button.grid(row=1,column=2,pady=10,padx=5)

tk.Label(master, text='Browse for seal (optional)  ',font=20).grid(row=2,column=0,pady=10,padx=5)
tk.Label(master, text='Browse for signature (optional)  ', font=20).grid(row=4,column=0,pady=10,padx=5)


def file1():
    global filename
    filename = filedialog.askopenfilename(initialdir='/', title='choose seal', filetypes=(('jpeg file','*.jpeg'),('all file','*.*')))
    my_label1 = tk.Label(master, text='', font=20)
    my_label1.grid(row=3, column=1)
    my_label1.configure(text=filename)

def file2():
    global filename1
    filename1 = filedialog.askopenfilename(initialdir='/', title='choose signature', filetypes=(('jpeg file','*.jpeg'),('all file','*.*')))
    my_label2 = tk.Label(master, text='', font=20)
    my_label2.grid(row=5, column=1)
    my_label2.configure(text=filename1)

my_button = tk.Button(master, text='Choose file', command=file1, font=20)
my_button.grid(row=2,column=1,pady=10,padx=5)
my_button = tk.Button(master, text='Choose file', command=file2, font=20)
my_button.grid(row=4,column=1,pady=10,padx=5)

def submit():
    str_1 = e3.get() 
    str_2 = e4.get()
    str1 = str_1.upper() 
    str2 = str_2.upper()
    if check_input(str1, str2):
        l = execute_range(str1, str2)
        messagebox.showinfo('Message info',f'Transcripts Generated, List of missing Roll nos: {l}')
    else:
        messagebox.showwarning('Message info','Enter a valid Range!')

my_button = tk.Button(master, text='Generate', command=submit, font=20)
my_button.grid(row=6,column=0,pady=10,padx=5)

def submit1():
    execute_all()

my_button = tk.Button(master, text='Generate  All', command=submit1, font=20)
my_button.grid(row=7,column=0,pady=10,padx=5)

master.mainloop()