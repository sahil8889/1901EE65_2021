import os
import csv
import openpyxl

os.mkdir('output')

with open('names-roll.csv', 'r') as csvfile:
    csvreader = csv.DictReader(csvfile)
    name = {}
    for row in csvreader:
        name[row['Roll']] = row['Name']

with open('subjects_master.csv', 'r') as csvfile:
    csvreader = csv.DictReader(csvfile)
    subject = {}
    for row in csvreader:
        list = []
        list.append(row['subname'])
        list.append(row['ltp'])
        subject[row['subno']] = list

with open('grades.csv', 'r') as csvfile:
    csvreader = csv.DictReader(csvfile)
    grades = {}
    for row in csvreader:
        list = []
        list.append(row['Roll'])
        list.append(row['Sem'])
        list.append(row['SubCode'])
        list.append(row['Credit'])
        list.append(row['Grade'])
        list.append(row['Sub_Type'])
        if(grades.get(row['Roll']) == None):
            grades[row['Roll']] = [list]
        else:
            grades[row['Roll']].append(list)

d = {}
for roll in grades:
    d1 = {}
    for row in grades[roll]:
        list = []
        list.append(row[2])
        list.append(subject[row[2]][0])
        list.append(subject[row[2]][1])
        list.append(row[3])
        list.append(row[5])
        list.append(row[4])
        if(d1.get(row[1]) == None):
            d1[row[1]] = [list]
        else:
            d1[row[1]].append(list)
    d[roll] = d1

for roll in grades:
    path = 'output/' + roll + '.xlsx'
    wb = openpyxl.Workbook()
    wb.create_sheet(index = 0, title='Overall')
    sheet = wb['Overall']
    sheet.append(['Name of Student', name[roll]])
    sheet.append(['Roll No.', roll])
    sheet.append(['Branch', roll[4:6]])
    
    GNE = {'AA':10,'AB':9,'BB':8,' BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,'DD*':4,'F*':0}
    semester = []
    spi = []
    cpi = []
    cred_taken = []
    total_cred_taken = []
    total_cred_sum = 0
    cpi_sum = 0
    for i in d[roll]:
        semester.append(i)
        spi_sum = 0
        cred_sum = 0
        for row in d[roll][i]:
            marks = float(GNE[row[5]])
            cred = float(row[3])
            spi_sum += marks*cred
            cred_sum += cred
        total_cred_sum += cred_sum
        cpi_sum += (spi_sum/cred_sum)*cred_sum
        spi.append(round(spi_sum/cred_sum, 2))
        cpi.append(round(cpi_sum/total_cred_sum, 2))
        cred_taken.append(cred_sum)
        total_cred_taken.append(total_cred_sum)
    sheet.append(['Semester No.'] + semester)
    sheet.append(['Semester wise Credit Taken'] + cred_taken)
    sheet.append(['SPI'] + spi)
    sheet.append(['Total Credits Taken'] + total_cred_taken)
    sheet.append(['CPI'] + cpi)

    for i in d[roll]:
        t = 'Sem' + i
        wb.create_sheet(index = int(i), title=t)
        sheet = wb[t]
        sheet.append(['Sl No.', 'Subject No.', 'Subject Name', 'L-T-P', 'Credit', 'Subject Type', 'Grade'])
        s = 0
        for row in d[roll][i]:
            s += 1
            row = [str(s)] + row
            sheet.append(row)

    wb.save(path)